# pip install openpyxl
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import os
import sys

applicationPath = os.path.dirname(sys.executable)

month = input('Introduce month: ')

inputPath = os.path.join(applicationPath, 'pivot_table.xlsx')
wb = load_workbook(inputPath)
sheet = wb['Report']

minColumn = wb.active.min_column
maxColumn = wb.active.max_column
minRow = wb.active.min_row
maxRow = wb.active.max_row

barchart = BarChart()

data = Reference(sheet,
                 min_col=minColumn+1,
                 max_col=maxColumn,
                 min_row=minRow,
                 max_row=maxRow)
categories = Reference(sheet,
                       min_col=minColumn,
                       max_col=minColumn,
                       min_row=minRow+1,
                       max_row=maxRow)

barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

sheet.add_chart(barchart, "B12")
barchart.title = 'Sales by Product line'
barchart.style = 5

for i in range(minColumn+1, maxColumn+1):
    letter = get_column_letter(i)
    sheet[f'{letter}{maxRow + 1}'] = f'=SUM({letter}{minRow + 1}:{letter}{maxRow})'
    sheet[f'{letter}{maxRow + 1}'].style = 'Currency'

sheet['A1'] = 'Sales Report'
sheet['A2'] = month
sheet['A1'].font = Font('Arial', bold=True, size=20)
sheet['A2'].font = Font('Arial', bold=True, size=10)

outputPath = os.path.join(applicationPath, f'report_{month}.xlsx')
wb.save(outputPath)
# pyinstaller --onefile excelReportExe.py